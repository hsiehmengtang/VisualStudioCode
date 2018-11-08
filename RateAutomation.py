################################################################################################################################
# Rate Automation
# 2018/06/19 Isaac Hsieh    This program autamatically updates RATE.xlsx file.
################################################################################################################################

import config
import numpy as np
import sys
import pandas
import openpyxl
import time
import datetime
import math
from dateutil.relativedelta import relativedelta
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from ftplib import FTP

################################################################################################################################
# Write rates into excel file
# currencyCode(str), year (int), month (int), clo (float), avg (float)
################################################################################################################################
def write_rates_into_file( currencyCode, year, month, clo, avg ):
    
    # Open file.
    fileName = config.RATE_PATH
    try:
        wb = openpyxl.load_workbook( fileName )
    except:
        print( "Error: Failed to open file " + fileName + '.' )
        sys.exit()

    # Write rates into file.
    ws = wb.active
    year = str( year )
    month = month_converter( str( month ) )
    date = year + '.' + month
    columnCurrencyCode = 2
    columnRateType = 3
    columnDate = 5
    columnRate = 6
    isCloRateExisted = False
    isAvgRateExisted = False

    # If the data is existed do update. Otherwise, add a new row.
    for rn in range( 1, ws.max_row+1 ):
        if ws.cell( row = rn, column = columnDate ).value == date:
            if ws.cell( row = rn, column = columnCurrencyCode ).value == currencyCode:
                if ws.cell( row = rn, column = columnRateType ).value == "CLO":
                    ws.cell( row = rn, column = columnRate ).value = clo
                    isCloRateExisted = True
                if ws.cell( row = rn, column = columnRateType ).value == "AVG":
                    ws.cell( row = rn, column = columnRate ).value = avg
                    isAvgRateExisted = True

    if isCloRateExisted == False:
        ws.append( ["ACTUAL", currencyCode, "CLO", "GLOBAL", date, clo] )
    if isAvgRateExisted == False:
        ws.append( ["ACTUAL", currencyCode, "AVG", "GLOBAL", date, avg] )

    # Save file
    wb.save( fileName )

################################################################################################################################
# Fetch rates from taiwan bank
# year (int), month (int), currencyCode(str)
################################################################################################################################
def fetch_rates_from_taiwan_bank( year, month, currencyCode ):
    
    clo = []
    avg = []
    isLastMonth = True

    # Currency string different from Taiwan bank and SAP.
    if currencyCode == "RMB":
        currencyCode = "CNY"

    # Get the average rate (YTD, from current month to January).
    year = str( year )
    for m in range( month, 0, -1 ):
        # Get data from Taiwan bank.
        mStr = month_converter( str( m ) )
        url = "http://rate.bot.com.tw/xrt/quote/" + year + "-" + mStr + "/" + currencyCode  
        try:
            html = pandas.read_html( url )
        except:
            print( "Error: Failed to open link " + url )
            sys.exit()
        currencyTable = html[0]
        
        # Rename columns.
        currencyTable.columns = [
            u"Date",
            u"Currency",
            u"CashBuying",
            u"CashSelling",
            u"SpotBuying",
            u"SpotSelling",
            u"NA"
        ]
        
        # If there is no data in this month, skip it.
        if math.isnan(currencyTable["CashBuying"][0]):
            continue

        # Retrieve all rates to calculate the average rate.
        if currencyTable["SpotBuying"][0] == '-':
            for element in currencyTable["CashBuying"]:
                avg.append(element)
            for element in currencyTable["CashSelling"]:
                avg.append(element)
        else:
            for element in currencyTable["SpotBuying"]:
                avg.append(element)
            for element in currencyTable["SpotSelling"]:
                avg.append(element)

        # Retrieve the latest rate, closing rate.
        if isLastMonth == True:
            isLastMonth = False
            if currencyTable["SpotBuying"][0] == '-':
                clo.append(currencyTable["CashBuying"][0])
                clo.append(currencyTable["CashSelling"][0])
            else:
                clo.append(currencyTable["SpotBuying"][0])
                clo.append(currencyTable["SpotSelling"][0])

    # Return closing and average rates.
    cloRate = np.mean(clo)
    avgRate = np.mean(avg)
    return round( cloRate, 5 ), round( avgRate, 5 )

################################################################################################################################
# Fetch rates from reuters
# currencyCode(str)
################################################################################################################################
def fetch_rates_from_reuters( currencyCode ):
    # Set path to chromedriver.exe.
    chrome_path = config.CHROME_PATH

    # Initial a web browser and create a browser object.
    browser = webdriver.Chrome( chrome_path )
    url = "https://www.reuters.com/finance/currencies"
    try:
        browser.get( url )
    except:
        print( "Error: Failed to open link " + url )
        sys.exit()

    # Select currencies.
    selectSrc = Select( browser.find_element_by_xpath( """//*[@id="currency-src-type"]""" ) )
    selectSrc.select_by_value( currencyCode )

    time.sleep( 2 )
    
    selectDest = Select( browser.find_element_by_xpath( """//*[@id="currency-dest-type"]""" ) )
    selectDest.select_by_value( "TWD" )

    time.sleep( 2 )

    # Click YTD.
    browser.find_element_by_xpath( """//*[@id="chartworksChart"]/div/div[1]/ul/li[1]""" ).click()
    browser.find_element_by_xpath( """//*[@id="chartworksChart"]/div/div[3]/div/ul/li[6]""" ).click()

    time.sleep( 2 )

    # Move the cursor to chartworks.
    chartworks = browser.find_element_by_xpath( """//*[@id="chartworksChart"]/div/div[2]""" )
    hover = ActionChains( browser ).move_to_element_with_offset( chartworks, 560, 70 )
    hover.perform()

    # Retrieve the price and date from current date to begining of the year.
    rates = []
    dates = []

    for _ in range( 1, 560 ):
        rate = browser.find_element_by_xpath( """//*[@id="chartworksChart"]/div/div[2]/div[1]/div[2]/div[2]/div[2]/span""" ).text
        rates.append( float(rate) )

        date = browser.find_element_by_xpath( """//*[@id="chartworksChart"]/div/div[2]/div[1]/div[2]/div[2]/div[1]""" ).text
        dates.append( date )

        hover = ActionChains( browser ).move_by_offset( -1, 0 )
        hover.perform()

    # Close the browser.
    browser.close()

    # Calculate closing rate and average rate.
    rateTable = pandas.DataFrame(
        {
            "Date": dates,
            "Rate": rates
        }
    )
    rateTable = rateTable.drop_duplicates( subset = "Date")
    
    # Return the closing and average rate.
    clo = rates[0]
    avg = round( rateTable["Rate"].mean(), 5 )
    return clo, avg

################################################################################################################################
# Month Converter
# month (str)
################################################################################################################################
def month_converter( month ):
    if( len( month ) == 1 ):
        month = '0' + month

    return month

################################################################################################################################
# Get File From FTP Server
# host (str), username (str), password (str), remoteDir (str), localDir (str), fileName (str)
################################################################################################################################
def get_file_from_ftpserver( host, username, password, remoteDir, localDir, fileName):
    # Connect to ftp server.
    ftp = FTP( host )
    ftp.login( user = username, passwd = password )
    ftp.cwd( remoteDir )

    # Get file from ftp server.
    localfile = open( localDir + fileName, "wb" )
    ftp.retrbinary( "RETR " + fileName, localfile.write, 1024 )
    ftp.quit()
    localfile.close()

################################################################################################################################
# Put File To FTP Server
# host (str), username (str), password (str), remoteDir (str), localDir (str), fileName (str)
################################################################################################################################
def put_file_to_ftpserver( host, username, password, remoteDir, localDir, fileName):
    # Connect to ftp server.
    ftp = FTP( host )
    ftp.login( user = username, passwd = password )
    ftp.cwd( remoteDir )

    # Put file to ftp server.
    localfile = open(localDir + fileName, "rb")
    ftp.storbinary("STOR " + fileName, localfile)
    ftp.quit()

################################################################################################################################
# Program starts here
################################################################################################################################
# FTP server configs.
host = config.HOST
username = config.USERNAME
password = config.PASSWORD
remoteDir = config.REMOTE_DIR
localDir = config.LOCAL_DIR
fileName = config.FILENAME

# Get RATE.xlsx from ftp server.
get_file_from_ftpserver( host, username, password, remoteDir, localDir, fileName)

# Get today's year and month.
year = datetime.date.today().year
month = datetime.date.today().month

# Fetch rates from taiwan bank and update them into RATE.xlsx.
currencyAvailableOnTaiwanBank = ["EUR", "JPY", "KRW", "MYR", "PHP", "RMB", "SGD", "USD"]
for currencyCode in currencyAvailableOnTaiwanBank:
    clo, avg = fetch_rates_from_taiwan_bank( year, month, currencyCode )
    write_rates_into_file( currencyCode, year, month, clo, avg )
    
# Fetch rates from reuters and update them into RATE.xlsx.
currencyCode = "DKK"
clo, avg = fetch_rates_from_reuters( currencyCode )
write_rates_into_file( currencyCode, year, month, clo, avg )

# Write TWD rate into RATE.xlsx.
currencyCode = "TWD"
write_rates_into_file( currencyCode, year, month, 1, 1 )

# Put RATE.xlsx to ftp server.
put_file_to_ftpserver( host, username, password, remoteDir, localDir, fileName)